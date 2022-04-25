import os
import sys
import time
import logging
import openpyxl
from datetime import date
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.common import exceptions as selenium_exception



class Payment():
    
    def __init__(self):

        chrome_driver_path = "chromedriver.exe"
        self.login_url = ""
        self.payment_url = ""
        
        self.login_credentials = {"username":"", "password":""}

        # self.excel_path = "Standard108180Pymt5219642.xlsx"
        self.excel_path = "New Microsoft Excel Worksheet.xlsx"
        starting_row = 1
        customer_column_number = 3
        invoice_column_number = 4
        invoice_amount_column_number = 9
        amount_paid_column_number = 11
        self.status_column_number = 14

        self.log_config()

        # initilize driver
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("useAutomationExtension", False)
        chrome_options.add_experimental_option("excludeSwitches",["enable-automation"])
        prefs = {"credentials_enable_service": False, "profile.password_manager_enabled": False}
        chrome_options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(executable_path=chrome_driver_path, options=chrome_options)
        self.driver.maximize_window()

        self.login() # redirect to login page
        
        # read excel
        try:
            wb_obj = openpyxl.load_workbook(self.excel_path)
            sheet_obj = wb_obj.active
            self.logger.debug(sheet_obj.max_row)
            
            for i in range(starting_row, sheet_obj.max_row + 1):
                try:
                    self.logger.info("Current excel row " + str(i))
                    customer = sheet_obj.cell(row = i, column = customer_column_number).value
                    invoice = sheet_obj.cell(row = i, column = invoice_column_number).value
                    invoice_amount = sheet_obj.cell(row = i, column = invoice_amount_column_number).value
                    amount_paid = sheet_obj.cell(row = i, column = amount_paid_column_number).value
                    if customer == "" or customer == None or invoice == None:
                        self.logger.info("Customer or invoice empty")
                        continue

                    self.logger.info("Customer " + customer + ", Invoice " + str(invoice))
                    if i == starting_row: # only for first row
                        self.payment(customer)
                    self.find_invoice(invoice, i, invoice_amount, amount_paid) 
                except Exception:
                    self.logger.exception("Exception occured while reading excel")
                    self.update_excel(i, "Error")
        except Exception:
            self.logger.exception("Unknown exception occured")
            self.terminate()

        self.logger.info("Precess complete")

        
    def log_config(self):
        
        try:
            log_dir = "logs"
            current_date = date.today().strftime("%d-%m-%Y")
            log_file = os.path.join(log_dir, current_date + ".log")
            
            if not os.path.isdir(log_dir):
                os.makedirs(log_dir)

            formatter = logging.Formatter("%(levelname)s:%(asctime)s:%(funcName)s:%(message)s")
            file_handler = logging.FileHandler(log_file)
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)

            stream_handler = logging.StreamHandler()
            stream_handler.setLevel(logging.INFO)
            
            self.logger=logging.getLogger(__name__)
            self.logger.setLevel(logging.DEBUG)
            self.logger.addHandler(file_handler)
            self.logger.addHandler(stream_handler)
        except Exception as ex:
            print(ex)
            sys.exit()

    def terminate(self):

        self.driver.quit()
        self.logger.info("Program exit")
        sys.exit()
    
    def wait_for_element(self, by, value):

        try:
            # self.logger.debug("Waiting for web element")
            time_end = time.time() + 60 * 5 # This will run for 5 min x 60 s = 300 seconds.
            while time.time() < time_end:
                try:
                    self.driver.find_element(by=by, value=value)
                    # self.logger.debug("Web element found")
                    break
                except selenium_exception.NoSuchElementException:
                    time.sleep(1)    
        except Exception:
            self.logger.exception("Exception occured while finding web elemnt")
    
    def login(self):
        
        try:
            username_input_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[3]/div/section[2]/div/form/fieldset/div[2]/input"
            password_input_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[7]/div/section[4]/div/div[1]/form/fieldset[4]/div/input"
            skip_button_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[14]/div/div/section/div/div[2]/form/div[2]/button"

            self.logger.info("Redirecting to login page")
            self.driver.get(self.login_url) # redirect to login page
            
            self.wait_for_element(By.XPATH, username_input_xpath)
            self.driver.find_element(by=By.XPATH, value=username_input_xpath).send_keys(self.login_credentials["username"]) # enter username
            self.driver.find_element(by=By.XPATH, value="/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[3]/div/section[2]/div/form/div[3]/button").click() # click continue button
            self.logger.debug("Continue with email")

            self.wait_for_element(By.XPATH, password_input_xpath)
            self.driver.find_element(by=By.XPATH, value=password_input_xpath).send_keys(self.login_credentials["password"]) # enter password
            self.driver.find_element(by=By.XPATH, value="/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[7]/div/section[4]/div/div[1]/form/div[3]/input").click() # click continue button
            self.logger.debug("Continue with password")
            
            self.wait_for_element(By.XPATH, skip_button_xpath)
            time.sleep(2)
            self.driver.find_element(by=By.XPATH, value=skip_button_xpath).click() # click on skip button
            self.logger.debug("Continue with skip for now")
            
            time.sleep(1)
            self.logger.info("Login successfull")

        except Exception:
            self.logger.exception("Exception occured while login")
            self.terminate()

    def payment(self, customer):

        self.driver.get(self.payment_url) # redirect to receive payment page

        choose_customer_dropdown_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[5]/div[5]/div[2]/div[1]/div[4]/div[2]/div[2]/div[1]/div/label/div/input"
        option_div_xpath = "/html/body/div[17]/div/div"
        
        self.wait_for_element(By.XPATH, choose_customer_dropdown_xpath)
        
        self.driver.find_element(by=By.XPATH, value=choose_customer_dropdown_xpath).clear()
        self.driver.find_element(by=By.XPATH, value=choose_customer_dropdown_xpath).send_keys(customer) # enter customer
        self.logger.debug("Entered customer")
        self.wait_for_element(By.XPATH, option_div_xpath)

        total_option = len(self.driver.find_element(by=By.XPATH, value=option_div_xpath).find_elements_by_xpath("./*")) # find total number of options
        self.logger.debug("Total option: " + str(total_option))
        for i in range(2, total_option + 1):
            option_xpath = "/html/body/div[17]/div/div/div[" + str(i) + "]"
            curr_customer = self.driver.find_element(By.XPATH, option_xpath + "/div/span[1]/span").text.strip()
            if curr_customer == customer:
                self.logger.debug("Customer found at : " + str(i))
                self.driver.find_element(by=By.XPATH, value=option_xpath).click() # click on customer
                self.logger.debug("Clicked on customer option")
                break
        time.sleep(3)
        
    def find_invoice(self, invoice_no, row_num, e_invoice_amount, e_amount_paid):

        try:
            invoice_input_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[2]/span[1]/div[1]/input"
            checkbox_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[1]/input"
            self.wait_for_element(By.XPATH, invoice_input_xpath)
            self.driver.find_element(by=By.XPATH, value=invoice_input_xpath).clear()
            self.driver.find_element(by=By.XPATH, value=invoice_input_xpath).send_keys(invoice_no) # enter invoice
            self.logger.debug("Entered invoice number")
            time.sleep(2)
            self.driver.find_element(by=By.XPATH, value= "//html").click() # false click to deselect invoice search input box
            
            try:
                self.driver.find_element(by=By.XPATH, value=checkbox_xpath) # check if checkbox element exists
                
                w_invoice_amount = self.driver.find_element(by=By.XPATH, value="/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[4]").text
                w_amount_paid = self.driver.find_element(by=By.XPATH, value="/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[5]").text
                
                if e_invoice_amount == e_amount_paid == int(float(w_invoice_amount)) == int(float(w_amount_paid)):
                    self.driver.find_element(by=By.XPATH, value=checkbox_xpath).click() # click on checkbox
                    self.logger.info("Checked")
                    self.update_excel(row_num, "Checked")
                else:
                    self.logger.info("Amount not matched, w_invoice_amount " + str(w_invoice_amount) + ", w_amount_paid " + str(w_amount_paid))
                    self.update_excel(row_num, "Not matched")
            except selenium_exception.NoSuchElementException:
                self.logger.info("Not found " + str(invoice_no) + ", row num " + str(row_num))
                self.update_excel(row_num, "Not found")
        except Exception:
            self.logger.exception("Exception occured in checking receive payment")
            self.update_excel(row_num, "Error")
    
    def update_excel(self, row_num, value):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            wordsheet = workbook.active
            wordsheet.cell(row=row_num,column=self.status_column_number).value = value
            workbook.save(self.excel_path)
            self.logger.info("Excel updated, row number " + str(row_num))
        except Exception:
            self.logger.exception("Exception occured while updating excel")
            self.terminate()


if __name__ == '__main__':
    _paymentObj = Payment()