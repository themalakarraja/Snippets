import os
import sys
import time
import logging
import openpyxl
from datetime import date
from selenium import webdriver 
from selenium.webdriver.common.by import By
from selenium.common import exceptions as selenium_exception



class Payment():
    def __init__(self):

        self.chrome_driver_path = "chromedriver.exe"
        self.login_url = ""
        self.receive_payment_url = ""
        
        self.login_credentials = {"username":"", "password":""}

        # self.excel_path = "Standard108180Pymt5219642.xlsx"
        self.excel_path = "new.xlsx"
        starting_row = 2
        customer_column_number = 3
        invoice_column_number = 4
        invoice_amount_column_number = 9
        amount_paid_column_number = 11
        self.status_column_number = 6

        self.log_config() # initilize log config
        self.initilize_driver() # initilize web driver
        self.login() # redirect to login page
        return
        # read excel
        try:
            wb_obj = openpyxl.load_workbook(self.excel_path)
            sheet_obj = wb_obj.active
            self.logger.debug(sheet_obj.max_row)
            
            for i in range(starting_row, sheet_obj.max_row + 1):
                try:
                    self.logger.info("Current excel row " + str(i))
                    customer = sheet_obj.cell(row = i, column = customer_column_number).value
                    invoice = sheet_obj.cell(row = i, column = invoice_column_number).value
                    invoice_amount = sheet_obj.cell(row = i, column = invoice_amount_column_number).value
                    amount_paid = sheet_obj.cell(row = i, column = amount_paid_column_number).value
                    if customer == "" or customer == None or invoice == None:
                        self.logger.warning("Customer or invoice empty")
                        continue

                    self.logger.debug("Customer " + customer + ", Invoice " + str(invoice))
                    self.receive_payment(customer, invoice, invoice_amount, amount_paid, i)
                except Exception:
                    self.logger.exception("Exception occured while reading excel")
                    self.update_excel(i, "Error")
        except Exception:
            self.logger.exception("Unknown exception occured")
            self.terminate()

        self.logger.info("Precess complete")

        
    def log_config(self):
        
        try:
            log_dir = "logs"
            current_date = date.today().strftime("%d-%m-%Y")
            log_file = os.path.join(log_dir, current_date + ".log")
            
            if not os.path.isdir(log_dir):
                os.makedirs(log_dir)

            formatter = logging.Formatter("%(levelname)s:%(asctime)s:%(funcName)s:%(message)s")
            file_handler = logging.FileHandler(log_file)
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)

            stream_handler = logging.StreamHandler()
            stream_handler.setLevel(logging.INFO)
            
            self.logger=logging.getLogger(__name__)
            self.logger.setLevel(logging.DEBUG)
            self.logger.addHandler(file_handler)
            self.logger.addHandler(stream_handler)
        except Exception as ex:
            print(ex)
            sys.exit()

    def initilize_driver(self):

        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option("useAutomationExtension", False)
        chrome_options.add_experimental_option("excludeSwitches",["enable-automation"])
        prefs = {"credentials_enable_service": False, "profile.password_manager_enabled": False}
        chrome_options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(executable_path=self.chrome_driver_path, options=chrome_options)
        self.driver.maximize_window()

    def terminate(self):

        self.driver.quit()
        self.logger.error("Program exit")
        sys.exit()
    
    def is_element_exists(self, by, value, wait_time = 0):
        '''
        This method waits until web element not found. Returns true if web element exists else return false
        '''
        try:
            if wait_time == 0:
                try:
                    self.driver.find_element(by=by, value=value)
                    return True
                except selenium_exception.NoSuchElementException:
                    return False
            else:
                time_end = time.time() + wait_time
                while time.time() < time_end:
                    try:
                        self.driver.find_element(by=by, value=value)
                        return True
                    except selenium_exception.NoSuchElementException:
                        time.sleep(1)
                self.logger.warning("Maximum time exceeded while finding web element")
                return False
        except Exception:
            self.logger.exception("Exception occured while finding web elemnt")
            return False
    
    def login(self):
        
        try:
            username_input_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[3]/div/section[2]/div/form/fieldset/div[2]/input"
            username_continue_button_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[3]/div/section[2]/div/form/div[3]/button"
            password_input_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[7]/div/section[4]/div/div[1]/form/fieldset[4]/div/input"
            password_continue_button_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[7]/div/section[4]/div/div[1]/form/div[3]/input"
            skip_button_xpath = "/html/body/div[2]/div/div[2]/table/tbody/tr/td[2]/div/div[2]/section[14]/div/div/section/div/div[2]/form/div[2]/button"

            self.logger.debug("Redirecting to login page")
            self.driver.get(self.login_url) # redirect to login page
            
             # enter username
            if not self.is_element_exists(By.XPATH, username_input_xpath, 60):
                raise Exception("Username input element xpath not found")
            self.driver.find_element(by=By.XPATH, value=username_input_xpath).send_keys(self.login_credentials["username"].strip())
            
             # click on continue button
            if not self.is_element_exists(By.XPATH, username_continue_button_xpath):
                raise Exception("Username continue button xpath not found")
            self.driver.find_element(by=By.XPATH, value=username_continue_button_xpath).click()
            self.logger.debug("Continue with email")

            # enter password
            if not self.is_element_exists(By.XPATH, password_input_xpath, 60):
                raise Exception("Password input element xpath not found")
            self.driver.find_element(by=By.XPATH, value=password_input_xpath).send_keys(self.login_credentials["password"].strip())

            # click on continue button
            if not self.is_element_exists(By.XPATH, password_continue_button_xpath):
                raise Exception("Password continue button xpath not found")
            self.driver.find_element(by=By.XPATH, value=password_continue_button_xpath).click()
            self.logger.debug("Continue with password")
            
            # click on skip button
            if not self.is_element_exists(By.XPATH, skip_button_xpath, 60):
                raise Exception("Skip button xpath not found")
            self.driver.find_element(by=By.XPATH, value=skip_button_xpath).click()
            self.logger.debug("Continue with skip for now")
            
            time.sleep(2)
            self.logger.info("Login successfull")
        except Exception:
            self.logger.exception("Exception occured while login")
            self.terminate()

    def receive_payment(self, customer, invoice_no, e_invoice_amount, e_amount_paid, row_num):
        
        try:
            self.driver.get(self.receive_payment_url) # redirect to receive payment page

            customer_dropdown_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[5]/div[5]/div[2]/div[1]/div[4]/div[2]/div[2]/div[1]/div/label/div/input"
            option_div_xpath = "/html/body/div[17]/div/div"
            
             # enter customer name
            if not self.is_element_exists(By.XPATH, customer_dropdown_xpath, 60):
                raise Exception("Customer dropdown xpath not found")
            self.driver.find_element(by=By.XPATH, value=customer_dropdown_xpath).clear()
            self.driver.find_element(by=By.XPATH, value=customer_dropdown_xpath).send_keys(str(customer).strip())
            self.logger.debug("Entered customer name")
            
            # select customer name from dropdown
            if not self.is_element_exists(By.XPATH, option_div_xpath, 60):
                raise Exception("Option div xpath not found")
            total_option = len(self.driver.find_element(by=By.XPATH, value=option_div_xpath).find_elements(by=By.XPATH, value="./*")) # find total number of options
            self.logger.debug("Total options in customer dropdown: " + str(total_option))
            for i in range(2, total_option + 1):
                option_xpath = "/html/body/div[17]/div/div/div[" + str(i) + "]"
                option_text_xpath = option_xpath + "/div/span[1]/span"
                
                if not self.is_element_exists(By.XPATH, option_xpath):
                    raise Exception("Option xpath not found")
                if not self.is_element_exists(By.XPATH, option_text_xpath):
                    raise Exception("Customer xpath not found")

                curr_customer = self.driver.find_element(By.XPATH, option_text_xpath).text.strip()
                if curr_customer == customer:
                    self.logger.debug("Customer found at: " + str(i))
                    self.driver.find_element(by=By.XPATH, value=option_xpath).click() # click on customer
                    self.logger.debug("Customer selected")
                    break
                else:
                    raise Exception("Customer not found")
            
            time.sleep(3)
            try:
                self.driver.find_element(By.XPATH, value="/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[5]/div[5]/div[2]/div[1]/div[8]/button").click() # click on Find by invoice no button
                self.driver.find_element(By.XPATH, value="/html/body/div[17]/div/div[1]/div/div/form/div/div/div/input").clear() # clear invoice no input field
                self.driver.find_element(By.XPATH, value="/html/body/div[17]/div/div[1]/div/div/form/div/div/div/input").send_keys(invoice_no) # enter invoice no
                self.driver.find_element(By.XPATH, value="/html/body/div[17]/div/div[1]/div/div/div/div[2]/button").click() # click in Find button
            except Exception:
                self.logger.exception("Exception occured while reading excel")
                self.update_excel(row_num, "Error")

        except Exception:
            self.logger.exception("Exception occured in receive payment")
            self.update_excel(row_num, "Error")

    
    def find_invoice(self, invoice_no, row_num, e_invoice_amount, e_amount_paid):

        try:
            invoice_input_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[2]/span[1]/div[1]/input"
            checkbox_xpath = "/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[1]/input"
            self.is_element_exists(By.XPATH, invoice_input_xpath)
            self.driver.find_element(by=By.XPATH, value=invoice_input_xpath).clear()
            self.driver.find_element(by=By.XPATH, value=invoice_input_xpath).send_keys(invoice_no) # enter invoice
            self.logger.debug("Entered invoice number")
            time.sleep(2)
            self.driver.find_element(by=By.XPATH, value= "//html").click() # false click to deselect invoice search input box
            
            try:
                self.driver.find_element(by=By.XPATH, value=checkbox_xpath) # check if checkbox element exists
                
                w_invoice_amount = self.driver.find_element(by=By.XPATH, value="/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[4]").text
                w_amount_paid = self.driver.find_element(by=By.XPATH, value="/html/body/div[3]/div/div[1]/div/div/div/div[1]/div[1]/div[7]/div[2]/div/div[2]/div[3]/div[2]/div/div/div[1]/div/div[4]/div/div/table/tr/td[5]").text
                
                if e_invoice_amount == e_amount_paid == int(float(w_invoice_amount)) == int(float(w_amount_paid)):
                    self.driver.find_element(by=By.XPATH, value=checkbox_xpath).click() # click on checkbox
                    self.logger.info("Checkbox checked")
                else:
                    self.logger.info("Amount not matched, w_invoice_amount " + str(w_invoice_amount) + ", w_amount_paid " + str(w_amount_paid))
            except selenium_exception.NoSuchElementException:
                self.logger.info("Not found " + str(invoice_no) + ", row num " + str(row_num))
                self.update_excel(row_num, "Not found")
        except Exception:
            self.logger.exception("Exception occured in checking receive payment")
            self.update_excel(row_num, "Error")
    
    def update_excel(self, row_num, value):
        try:
            workbook = openpyxl.load_workbook(self.excel_path)
            wordsheet = workbook.active
            wordsheet.cell(row=row_num,column=self.status_column_number).value = value
            workbook.save(self.excel_path)
            self.logger.info("Excel updated, row number " + str(row_num))
        except Exception:
            self.logger.exception("Exception occured while updating excel")
            self.terminate()


if __name__ == '__main__':
    _paymentObj = Payment()