from selenium import webdriver 
from selenium.webdriver.common.by import By
import time
import openpyxl
from bs4 import BeautifulSoup

class Inventory:
    
    def __init__(self):

        chrome_driver_path = "chromedriver.exe"
        self.login_url = "https://www.hitpromo.net/site/login"
        self.product_url = "https://www.hitpromo.net/product/show/"
        self.login_credentials = {"username":"", "password":""}
        self.is_login = False
        self.driver = webdriver.Chrome(executable_path=chrome_driver_path)

    def __login(self):
        
        self.driver.maximize_window()
        print("Redirected to login..")
        self.driver.get(self.login_url)
        self.driver.find_element(by=By.ID, value="LoginForm_username").send_keys(self.login_credentials["username"])
        self.driver.find_element(by=By.ID, value="LoginForm_password").send_keys(self.login_credentials["password"])
        self.driver.find_element(by=By.NAME, value="yt0").click()
        self.is_login = True
        print("Login completed..")


    def getInventory(self, style_no):
        
        if not self.is_login:
            self.__login()
        
        print("Redirected to product page.. style_no: " + str(style_no))
        self.driver.get(self.product_url + str(style_no))

        print("Waiting for loading inventory..")
        result = "Loading Inventory"
        #since inventory data loads through ajax in website, it will check in every one second if data loads
        while(result in "Loading Inventory" or result in "LOADING INVENTORY"):
            time.sleep(1)
            try:
                # result = self.driver.find_element_by_id("inventory").text
                result = self.driver.find_element(by=By.ID, value="inventory").text
            except:
                pass
        print("Loading inventory completed..")
        element = self.driver.find_element(by=By.XPATH, value= '//*[@id="inventoryTable"]')
        element = element.get_attribute('innerHTML')
        
        return self.__formatData(element)

    def __formatData(self, inventory_str):
        """
        This method takes imput and clean the data 
        and returns color name with quantity in csv format.
        
        return format
        Blue,100
        Red,50
        Green,2005
        """

        print("Start converting html to csv ..")
        inventory = ""
        soup = BeautifulSoup(inventory_str, 'lxml')
        trs = soup.find("tbody").find_all("tr")
        for tr in trs:
            try:
                tds = tr.find_all("td")
                color_name = tds[0].text
                quantity = tds[1].find("strong").text
                inventory += color_name + "," + quantity + "\n"
            except:
                pass
        print("Converting html to csv completed..")
        return inventory

    def closeBrowser(self):
        self.driver.close()



def findColorInCSV(csv_str, color):
    """
    This method takes inventory csv string as inpuet and 
    returns the quantity of the provided color as parameter
    """
    print("Finding color.. " + color)
    csv_str = csv_str.upper()
    color = color.upper()
    csv_str = csv_str.split("\n")
    for row in csv_str:
        words = row.split(",")
        if color in words[0]:
            return words[1]
    return None

def updateXlsx(excel_path, row_num, column_num, value):

    workbook = openpyxl.load_workbook(excel_path)
    wordsheet = workbook.active
    wordsheet.cell(row=row_num,column=column_num).value = value
    workbook.save(excel_path)
    print("Updated excel..")


def start():

    excel_path = "TestWorksheet.xlsx"
    inventory = ""
    previous_style = ""
    starting_row = 2
    quantity_row_num = 9
    inventoryObj = Inventory()

    wb_obj = openpyxl.load_workbook(excel_path)
    sheet_obj = wb_obj.active
    for i in range(starting_row, sheet_obj.max_row + 1):
        color_name = sheet_obj.cell(row = i, column = 3).value
        style = sheet_obj.cell(row = i, column = 6).value
        if style != previous_style:
            inventory = inventoryObj.getInventory(style)
        quantity = findColorInCSV(inventory, color_name.split()[0])
        if quantity is not None:
            updateXlsx(excel_path, i, quantity_row_num, quantity)
        previous_style = style
    print("Process completed..")
    inventoryObj.closeBrowser()

start()
